"""
XLSX parser for GDPdU/GoBD dossier spreadsheets.

Handles German number formats (comma decimal), German date formats (DD.MM.YYYY),
and maps known filenames to RecordType and entity extraction strategies.
"""

from __future__ import annotations

import logging
import re
import uuid
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.normalization.models import (
    EntityRef,
    NormalizedRecord,
    RecordType,
    SourceProvenance,
)

logger = logging.getLogger(__name__)

NAMESPACE_URL = uuid.NAMESPACE_URL

# Filename pattern -> (RecordType, entity extraction strategy)
_FILENAME_TYPE_MAP: list[tuple[str, RecordType, str]] = [
    ("Berechtigungsauswertung", RecordType.permission, "user"),
    ("OP-Liste_Debitoren", RecordType.open_item, "customer"),
    ("OP-Liste_Kreditoren", RecordType.open_item, "vendor"),
    ("OP-Liste", RecordType.open_item, "vendor"),
    ("Saldenliste", RecordType.balance, "account"),
    ("Abstimmung", RecordType.balance, "account"),
]

# Column name patterns for entity extraction
_ACCOUNT_COLUMNS = re.compile(
    r"(konto|account|sachkonto|kto)", re.IGNORECASE
)
_VENDOR_COLUMNS = re.compile(
    r"(lieferant|kreditor|vendor|supplier)", re.IGNORECASE
)
_CUSTOMER_COLUMNS = re.compile(
    r"(kunde|debitor|customer|client)", re.IGNORECASE
)
_USER_COLUMNS = re.compile(
    r"(benutzer|user|anwender|erfasser|bearbeiter)", re.IGNORECASE
)
_AMOUNT_COLUMNS = re.compile(
    r"(betrag|saldo|amount|summe|wert|haben|soll|debit|credit)", re.IGNORECASE
)
_DATE_COLUMNS = re.compile(
    r"(datum|date|buchungsdatum|belegdatum|faelligkeit)", re.IGNORECASE
)

# German date pattern DD.MM.YYYY
_GERMAN_DATE_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$")
# German number with comma decimal: 1.234,56 or just 1234,56
_GERMAN_NUMBER_RE = re.compile(r"^-?\d{1,3}(?:\.\d{3})*,\d{1,2}$")

# Header detection bounds - see _detect_header_row for the full heuristic.
_MAX_HEADER_SCAN_ROWS = 15
_MIN_HEADER_NON_EMPTY_CELLS = 2
_MIN_HEADER_FILL_RATIO = 0.5
_MIN_HEADER_TEXT_RATIO = 0.8
_MIN_HEADER_UNIQUE_RATIO = 0.8


def _make_record_id(dossier_id: str, file_id: str, row: int, index: int) -> str:
    """Generate stable UUID5 for a record."""
    dossier_ns = uuid.uuid5(NAMESPACE_URL, f"dossier:{dossier_id}")
    return str(uuid.uuid5(dossier_ns, f"{file_id}:{row}:{index}"))


def _detect_record_type(filename: str) -> tuple[RecordType, str]:
    """Determine RecordType and entity strategy from filename."""
    for pattern, record_type, entity_strategy in _FILENAME_TYPE_MAP:
        if pattern in filename:
            return record_type, entity_strategy
    # Default fallback
    return RecordType.balance, "account"


def _parse_german_date(value: str) -> str | None:
    """Convert DD.MM.YYYY to ISO 8601 date string."""
    if not isinstance(value, str):
        return None
    match = _GERMAN_DATE_RE.match(value.strip())
    if match:
        day, month, year = match.groups()
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return None


def _parse_german_number(value: str) -> float | None:
    """
    Convert German number format to float.

    Examples:
        "1.234,56" -> 1234.56
        "-1.234,56" -> -1234.56
        "1234,56" -> 1234.56
    """
    if not isinstance(value, str):
        return None
    cleaned = value.strip()
    if _GERMAN_NUMBER_RE.match(cleaned):
        # Remove thousand separators (dots), replace comma with dot
        cleaned = cleaned.replace(".", "").replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None
    # Try simple comma decimal without thousand separator
    if "," in cleaned and "." not in cleaned:
        try:
            return float(cleaned.replace(",", "."))
        except ValueError:
            return None
    return None


def _normalize_cell_value(value) -> str | int | float | None:
    """Normalize a cell value to a JSON-compatible type."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        # Try German number
        german_num = _parse_german_number(stripped)
        if german_num is not None:
            return german_num
        # Try plain number
        try:
            if "." in stripped and "," not in stripped:
                return float(stripped)
            if stripped.isdigit() or (
                stripped.startswith("-") and stripped[1:].isdigit()
            ):
                return int(stripped)
        except (ValueError, IndexError):
            pass
        return stripped
    # datetime objects
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _detect_header_row(rows: list[tuple]) -> int | None:
    """
    Find the real header row within the first rows of a sheet.

    Real-world GDPdU/GoBD exports routinely stack a company/report banner
    line and a blank spacer row above the actual column header, so `rows[0]`
    is almost never the header. This scans only the first
    `_MAX_HEADER_SCAN_ROWS` rows (bounded - it never reads the whole sheet)
    and returns the 0-based index of the first row that looks like a header,
    or `None` if no row qualifies.

    A row qualifies as a header when all of these hold:
      - it has at least `_MIN_HEADER_NON_EMPTY_CELLS` non-empty cells, which
        rules out single-cell banner/section-label rows and blank spacers;
      - its fill ratio (non-empty cells / total columns) is at least
        `_MIN_HEADER_FILL_RATIO` - header rows are fully populated, banners
        are one cell wide in a many-column sheet;
      - its text ratio (non-empty cells that normalize to strings rather
        than numbers or dates) is at least `_MIN_HEADER_TEXT_RATIO` - column
        names are labels, not data values; this is what rejects data rows
        that happen to be dense and mostly-unique;
      - its uniqueness ratio (distinct values / non-empty cells) is at least
        `_MIN_HEADER_UNIQUE_RATIO` - column names don't repeat within a row.

    The first qualifying row is returned rather than the best-scoring one:
    a sheet has one header, and it always precedes the data it describes, so
    the earliest match in scan order is correct by construction once the
    thresholds above have filtered out banners and spacers.

    Failure mode: sheets with no row meeting all four thresholds in the
    scanned window return `None`. This is a deliberate outcome, not an
    error - it covers label/value reconciliation sheets that have no
    tabular header at all (e.g. two-column "metric -> amount" sheets), and
    also covers headers that start deeper than `_MAX_HEADER_SCAN_ROWS`, which
    this heuristic does not support. Callers must fall back to positional
    `column_N` names in both cases and must not run entity-extraction
    heuristics that assume a real header column name.
    """
    scan_limit = min(len(rows), _MAX_HEADER_SCAN_ROWS)
    for idx in range(scan_limit):
        row = rows[idx]
        total_columns = len(row)
        if total_columns == 0:
            continue

        non_empty_values = [
            value
            for value in (_normalize_cell_value(cell) for cell in row)
            if value is not None and value != ""
        ]
        if len(non_empty_values) < _MIN_HEADER_NON_EMPTY_CELLS:
            continue

        fill_ratio = len(non_empty_values) / total_columns
        if fill_ratio < _MIN_HEADER_FILL_RATIO:
            continue

        text_count = sum(1 for value in non_empty_values if isinstance(value, str))
        text_ratio = text_count / len(non_empty_values)
        if text_ratio < _MIN_HEADER_TEXT_RATIO:
            continue

        unique_count = len({str(value) for value in non_empty_values})
        unique_ratio = unique_count / len(non_empty_values)
        if unique_ratio < _MIN_HEADER_UNIQUE_RATIO:
            continue

        return idx

    return None


def _extract_entities(
    row_data: dict[str, str | int | float | None],
    headers: list[str],
    entity_strategy: str | None,
) -> list[EntityRef]:
    """
    Extract entity references from a row based on column names.

    `entity_strategy` is `None` when no real header row was detected for the
    sheet: without genuine column names, `headers[0]` is a synthetic
    `column_0` placeholder, and the positional "first column is the entity
    id" fallback below would misfire on arbitrary label text. Column-name
    pattern matches still run in that case but naturally find nothing, since
    patterns like `_ACCOUNT_COLUMNS` never match `column_N`.
    """
    entities: list[EntityRef] = []
    seen: set[tuple[str, str]] = set()

    for col_name, cell_value in row_data.items():
        if cell_value is None or cell_value == "":
            continue
        str_value = str(cell_value).strip()
        if not str_value:
            continue

        entity_type: str | None = None
        if _ACCOUNT_COLUMNS.search(col_name):
            entity_type = "account"
        elif _VENDOR_COLUMNS.search(col_name):
            entity_type = "vendor"
        elif _CUSTOMER_COLUMNS.search(col_name):
            entity_type = "customer"
        elif _USER_COLUMNS.search(col_name):
            entity_type = "user"

        if entity_type is None and entity_strategy in (
            "account",
            "vendor",
            "customer",
            "user",
        ):
            # For the primary entity column (first column that looks like an ID)
            if col_name == headers[0] and re.match(r"^[\w\-]+$", str_value):
                entity_type = entity_strategy

        if entity_type and (entity_type, str_value) not in seen:
            seen.add((entity_type, str_value))
            entities.append(
                EntityRef(
                    entity_type=entity_type,
                    entity_id=str_value,
                    label=str_value,
                )
            )

    return entities


def _extract_amount(
    row_data: dict[str, str | int | float | None],
) -> float | None:
    """Find the primary amount in a row."""
    for col_name, value in row_data.items():
        if _AMOUNT_COLUMNS.search(col_name) and isinstance(value, (int, float)):
            return float(value)
    return None


def _extract_date(
    row_data: dict[str, str | int | float | None],
) -> str | None:
    """Find the primary date in a row."""
    for col_name, value in row_data.items():
        if _DATE_COLUMNS.search(col_name) and isinstance(value, str):
            iso_date = _parse_german_date(value)
            if iso_date:
                return iso_date
            # Already ISO format?
            if re.match(r"^\d{4}-\d{2}-\d{2}", value):
                return value[:10]
    return None


def parse_xlsx_file(
    file_path: Path,
    relative_path: str,
    dossier_id: str,
    file_id: str,
) -> list[NormalizedRecord]:
    """
    Parse an XLSX file into NormalizedRecords.

    Reads all sheets, detects the real header row per sheet (see
    `_detect_header_row` - GDPdU/GoBD exports routinely have a banner and a
    blank row above the actual header), and converts each row below the
    header into a NormalizedRecord with entity extraction and German format
    handling. Sheets where no header is found (e.g. label/value
    reconciliation sheets) fall back to positional `column_N` names and
    treat every row as data.

    Args:
        file_path: Absolute path to the XLSX file.
        relative_path: Path relative to dossier root.
        dossier_id: ID of the parent dossier.
        file_id: Stable ID of this file in the manifest.

    Returns:
        List of NormalizedRecords. Empty list on failure.
    """
    records: list[NormalizedRecord] = []
    filename = Path(relative_path).stem

    record_type, entity_strategy = _detect_record_type(filename)

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except (InvalidFileException, Exception) as exc:
        logger.warning(
            "Failed to open XLSX file %s: %s", relative_path, exc
        )
        return []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                logger.debug(
                    "Sheet '%s' in %s is empty, skipping.",
                    sheet_name,
                    relative_path,
                )
                continue

            header_row_idx = _detect_header_row(rows)

            if header_row_idx is not None:
                raw_headers = rows[header_row_idx]
                headers: list[str] = []
                for i, h in enumerate(raw_headers):
                    if h is not None:
                        headers.append(str(h).strip())
                    else:
                        headers.append(f"column_{i}")
                data_rows = enumerate(
                    rows[header_row_idx + 1 :], start=header_row_idx + 2
                )
            else:
                logger.info(
                    "No header row detected in sheet '%s' of %s within the "
                    "first %d rows; using positional column names and "
                    "treating every row as data.",
                    sheet_name,
                    relative_path,
                    _MAX_HEADER_SCAN_ROWS,
                )
                headers = [f"column_{i}" for i in range(len(rows[0]))]
                data_rows = enumerate(rows, start=1)

            # A synthetic header carries no real column names, so the
            # first-column entity-id fallback in _extract_entities must not
            # run - see that function's docstring.
            row_entity_strategy = (
                entity_strategy if header_row_idx is not None else None
            )

            # Process data rows
            for row_idx, row in data_rows:
                # Skip completely empty rows
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                try:
                    # Build data dict
                    row_data: dict[str, str | int | float | None] = {}
                    for col_idx, cell_value in enumerate(row):
                        if col_idx < len(headers):
                            col_name = headers[col_idx]
                        else:
                            col_name = f"column_{col_idx}"
                        row_data[col_name] = _normalize_cell_value(cell_value)

                    # Extract entities
                    entities = _extract_entities(
                        row_data, headers, row_entity_strategy
                    )

                    # Extract amount and date
                    amount = _extract_amount(row_data)
                    date = _extract_date(row_data)

                    # Build relationships dict from entities
                    relationships: dict[str, str] = {}
                    for entity in entities:
                        relationships[entity.entity_type] = entity.entity_id

                    record_id = _make_record_id(
                        dossier_id, file_id, row_idx, 0
                    )

                    record = NormalizedRecord(
                        record_id=record_id,
                        dossier_id=dossier_id,
                        record_type=record_type,
                        source=SourceProvenance(
                            file_id=file_id,
                            relative_path=relative_path,
                            row_number=row_idx,
                            sheet=sheet_name,
                            columns=headers,
                        ),
                        entities=entities,
                        relationships=relationships,
                        data=row_data,
                        date=date,
                        amount=amount,
                        currency="EUR" if amount is not None else None,
                    )
                    records.append(record)

                except Exception as exc:
                    logger.warning(
                        "Skipping row %d in sheet '%s' of %s: %s",
                        row_idx,
                        sheet_name,
                        relative_path,
                        exc,
                    )
                    continue

    finally:
        wb.close()

    logger.info(
        "Parsed %d records from XLSX %s", len(records), relative_path
    )
    return records
